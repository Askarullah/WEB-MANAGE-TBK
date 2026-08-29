# app.py - Main Flask application
from flask import Flask, render_template, request, jsonify, session, send_file
from flask_cors import CORS
import pandas as pd
import sys
import os
import webbrowser
import threading
import time
import uuid
from werkzeug.utils import secure_filename
# Removed unused tempfile import
import json
from datetime import datetime
import io
import openpyxl
import paramiko
import re
import socket

import dotenv
from dotenv import load_dotenv

load_dotenv()

# Use os.path.join for better cross-platform compatibility
app = Flask(__name__, template_folder='templates')
CORS(app)  # Enable CORS for all routes
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'your-secret-key-change-this')  # Change this in production
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 52428800))  # 50MB max file size

# Global storage for workbook data (in production, use Redis or database)
workbook_storage = {}

# MikroTik connection settings
MIKROTIK_SSH_PORT = int(os.getenv('MIKROTIK_SSH_PORT'))
MIKROTIK_USERNAME = os.getenv('MIKROTIK_USERNAME')
MIKROTIK_PASSWORD = os.getenv('MIKROTIK_PASSWORD')

OLT_SSH_PORT = int(os.getenv('OLT_SSH_PORT', '22'))
OLT_USERNAME = os.getenv('OLT_USERNAME')
OLT_PASSWORD = os.getenv('OLT_PASSWORD')

# ============================================================================
# MikroTik Authentication and SSH Connection Management
# ============================================================================

def load_mikrotik_credentials():
    """
    Load MikroTik credentials from environment variables
    Returns list of credentials with actual passwords for internal use
    """
    credentials = []
    
    # Read all available users (up to 10 for safety)
    for i in range(1, 11):
        if i == 1:
            username_key = 'MIKROTIK_USERNAME'
            password_key = 'MIKROTIK_PASSWORD'
        else:
            username_key = f'MIKROTIK_USERNAME_{i}'
            password_key = f'MIKROTIK_PASSWORD_{i}'
        
        username = os.getenv(username_key)
        password = os.getenv(password_key)
        
        if username:
            credentials.append({
                'username': username,
                'password': password if password and password.strip() else None
            })
    
    return credentials

def get_stored_password_for_user(username):
    """
    Retrieve stored password for a specific username from environment variables.
    This function supports multiple username/password combinations stored in
    the .env file (MIKROTIK_USERNAME_1 through MIKROTIK_USERNAME_10).
    
    Args:
        username (str): The username to find the password for
        
    Returns:
        str or None: The password if found, None otherwise
    """
    # Check all possible username/password combinations (up to 10 sets)
    for i in range(1, 11):
        if i == 1:
            # Check primary credentials (without number suffix)
            env_username = os.getenv('MIKROTIK_USERNAME')
            env_password = os.getenv('MIKROTIK_PASSWORD')
        else:
            # Check numbered credentials (MIKROTIK_USERNAME_2, etc.)
            env_username = os.getenv(f'MIKROTIK_USERNAME_{i}')
            env_password = os.getenv(f'MIKROTIK_PASSWORD_{i}')
        
        # Return password if username matches and password exists
        if env_username == username and env_password:
            return env_password
    
    return None

def connect_to_mikrotik(ip, command, username=None, password=None):
    """
    Establish SSH connection to MikroTik router and execute commands.
    This function handles authentication, connection management, command execution,
    and comprehensive error handling for MikroTik RouterOS devices.
    
    Args:
        ip (str): IP address of the MikroTik device
        command (str): RouterOS command to execute
        username (str, optional): Username for authentication (uses default if None)
        password (str, optional): Password for authentication (uses default if None)
                                 Use 'stored' to lookup password from environment
        
    Returns:
        tuple: (success: bool, result: str) - Success status and command output or error message
    """
    ssh_client = None
    
    # Determine authentication credentials
    auth_username = username if username else MIKROTIK_USERNAME
    auth_password = password if password else MIKROTIK_PASSWORD
    
    # Handle special case: lookup stored password for username
    if password == 'stored':
        auth_password = get_stored_password_for_user(username)
        if not auth_password:
            return False, f"No stored password found for user {username}"
    
    try:
        # Initialize SSH client with security settings
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())  # Accept unknown host keys
        ssh_client.load_system_host_keys()  # Load known host keys
        
        print(f"Attempting to connect to {ip}:{MIKROTIK_SSH_PORT} as {auth_username}")
        
        # Establish SSH connection with comprehensive timeout settings
        ssh_client.connect(
            hostname=ip,
            username=auth_username,
            password=auth_password,
            port=MIKROTIK_SSH_PORT,
            timeout=30,  # Connection timeout (30 seconds)
            auth_timeout=30,  # Authentication timeout (30 seconds)
            banner_timeout=30,  # Banner exchange timeout (30 seconds)
            look_for_keys=False,  # Don't search for SSH private keys
            allow_agent=False,  # Don't use SSH agent
            compress=False  # Disable compression for better compatibility
        )
        
        print(f"Successfully connected to {ip}")
        
        # Execute RouterOS command with timeout
        stdin, stdout, stderr = ssh_client.exec_command(command, timeout=30)
        
        # Wait for command completion and get exit status
        exit_status = stdout.channel.recv_exit_status()
        
        # Read command output and error streams
        output = stdout.read().decode('utf-8', errors='ignore')
        error = stderr.read().decode('utf-8', errors='ignore')
        
        print(f"Command executed. Exit status: {exit_status}")
        print(f"Output length: {len(output)} characters")
        
        # Check for command execution errors
        if exit_status != 0:
            return False, f"Command failed with exit status {exit_status}: {error}"
        
        # Check for RouterOS syntax errors
        if error and "syntax error" in error.lower():
            return False, f"MikroTik syntax error: {error}"
        
        return True, output
        
    except paramiko.AuthenticationException as e:
        # Handle authentication failures
        error_msg = f"Authentication failed for {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except paramiko.SSHException as e:
        # Handle SSH protocol errors
        error_msg = f"SSH connection failed to {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except socket.timeout:
        # Handle connection timeouts
        error_msg = f"Connection timeout to {ip}:{MIKROTIK_SSH_PORT}"
        print(error_msg)
        return False, error_msg
        
    except socket.error as e:
        # Handle network connectivity errors
        error_msg = f"Network error connecting to {ip}:{MIKROTIK_SSH_PORT}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except Exception as e:
        # Handle any other unexpected errors
        error_msg = f"Unexpected error connecting to {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    finally:
        # Ensure SSH connection is always closed
        if ssh_client:
            try:
                ssh_client.close()
                print(f"SSH connection to {ip} closed")
            except:
                pass

def connect_to_olt(ip, command, username=None, password=None):
    """
    Connect to OLT device via Telnet with full login sequence using raw sockets
    Handles pagination (--More-- prompts) automatically
    """
    sock = None
    auth_username = username if username else OLT_USERNAME
    auth_password = password if password else OLT_PASSWORD

    if not auth_username or not auth_password:
        return False, "OLT credentials are not configured"

    try:
        print(f"Connecting to OLT {ip} via Telnet...")
        
        # Step 1: Create socket connection
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(30)
        sock.connect((ip, OLT_SSH_PORT))
        
        print(f"Telnet connected to {ip}, starting login sequence...")
        
        output = ""
        command_output = ""
        
        # Helper function to receive data
        def receive_data(timeout=5):
            sock.settimeout(timeout)
            data = b""
            try:
                while True:
                    chunk = sock.recv(4096)
                    if not chunk:
                        break
                    data += chunk
                    time.sleep(0.1)
                    # Check if there's more data
                    sock.settimeout(0.1)
            except socket.timeout:
                pass
            except Exception as e:
                print(f"Receive error: {e}")
            sock.settimeout(timeout)
            return data.decode('utf-8', errors='ignore')
        
        # Helper function to send data
        def send_data(text):
            sock.sendall((text + "\r\n").encode('utf-8'))
            time.sleep(0.5)
        
        # Step 2: Wait for initial banner
        time.sleep(2)
        initial_banner = receive_data(timeout=3)
        output += initial_banner
        print(f"Initial banner: {repr(initial_banner[:200])}")
        
        # Step 3: Wait for "Login:" or "Username:" prompt
        print("Waiting for login prompt...")
        time.sleep(1)
        login_prompt = receive_data(timeout=5)
        output += login_prompt
        print(f"Login prompt: {repr(login_prompt[:200])}")
        
        # Step 4: Send username
        print(f"Sending username: {auth_username}")
        send_data(auth_username)
        
        # Step 5: Wait for "Password:" prompt
        print("Waiting for password prompt...")
        time.sleep(1)
        password_prompt = receive_data(timeout=5)
        output += password_prompt
        print(f"Password prompt: {repr(password_prompt[:200])}")
        
        # Step 6: Send password
        print("Sending password")
        send_data(auth_password)
        time.sleep(2)
        
        # Step 7: Read response after login
        login_response = receive_data(timeout=3)
        output += login_response
        print(f"Login response: {repr(login_response[:200])}")
        
        # Check if login failed
        if "incorrect" in login_response.lower() or "failed" in login_response.lower() or "denied" in login_response.lower():
            return False, "Authentication failed - incorrect username or password"
        
        # Step 8: Wait for user mode prompt ">"
        print("Waiting for user prompt '>'...")
        time.sleep(1)
        user_prompt = receive_data(timeout=2)
        output += user_prompt
        print(f"User prompt: {repr(user_prompt[:100])}")
        
        # Step 9: Send "enable" command to enter privileged mode
        print("Sending 'enable' command...")
        send_data('enable')
        time.sleep(1)
        
        # Step 10: Wait for enable password prompt
        print("Waiting for enable password prompt...")
        enable_prompt = receive_data(timeout=3)
        output += enable_prompt
        print(f"Enable prompt: {repr(enable_prompt[:100])}")
        
        # Step 11: Send enable password
        print("Sending enable password")
        send_data(auth_password)
        time.sleep(2)
        
        # Step 12: Read response after enable
        enable_response = receive_data(timeout=3)
        output += enable_response
        print(f"Enable response: {repr(enable_response[:200])}")
        
        # Step 13: Verify we're in enabled mode (should see "#" prompt)
        if '#' not in enable_response and '#' not in output:
            return False, "Failed to enter enabled mode. Check credentials."
        
        print("Successfully entered enabled mode!")
        
        # Step 14: Send the actual command
        print(f"Sending command: {command}")
        send_data(command)
        time.sleep(4)  # Wait for command execution
        
        # Step 15: Read command output with pagination handling
        print("Reading command output (handling pagination)...")
        command_output = ""
        max_pages = 100  # Safety limit to prevent infinite loops
        page_count = 0
        
        while page_count < max_pages:
            # Read available data
            chunk = receive_data(timeout=5)
            command_output += chunk
            
            # Check for pagination prompts
            if '--More--' in chunk or '-- More --' in chunk or 'More' in chunk:
                print(f"Page {page_count + 1}: Found pagination prompt, sending space...")
                # Send space to continue (space key is better than Enter for pagination)
                sock.sendall(b" ")
                time.sleep(1)
                page_count += 1
            else:
                # No more pagination, break the loop
                print("No more pagination found")
                break
        
        # Wait a bit more and read any remaining output
        time.sleep(2)
        additional_output = receive_data(timeout=2)
        command_output += additional_output
        
        print(f"Command output length: {len(command_output)} characters")
        print(f"Pages processed: {page_count}")
        
        # Step 16: Send exit command
        print("Sending exit command...")
        send_data('exit')
        time.sleep(0.5)
        
        # Get any remaining output
        final_output = receive_data(timeout=1)
        command_output += final_output
        
        if not command_output.strip():
            return False, "No output received from command"
        
        # Clean output - remove command echo, prompts, and pagination artifacts
        lines = command_output.split('\n')
        clean_lines = []
        
        for line in lines:
            line_lower = line.lower().strip()
            # Skip lines containing the command itself, prompts, pagination, or login text
            if (command.lower() in line_lower or 
                line_lower.endswith('>') or 
                line_lower.endswith('#') or
                'login:' in line_lower or
                'username:' in line_lower or
                'password:' in line_lower or
                '--more--' in line_lower or
                '-- more --' in line_lower or
                line_lower == 'enable' or
                line_lower == 'exit'):
                continue
            
            # Remove ANSI escape codes (used for clearing --More-- prompts)
            line_clean = line.replace('\x1b[K', '').replace('\x1b[?7h', '').replace('\x1b[?7l', '')
            line_clean = line_clean.replace('\r', '').strip()
            
            if line_clean:  # Only include non-empty lines
                clean_lines.append(line_clean)
        
        clean_output = '\n'.join(clean_lines).strip()
        
        print(f"Command executed successfully on {ip}")
        return True, clean_output if clean_output else command_output.strip()

    except socket.timeout:
        error_msg = f"Connection timeout to {ip}:{OLT_SSH_PORT}"
        print(error_msg)
        return False, error_msg

    except socket.error as e:
        error_msg = f"Socket error connecting to {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg

    except Exception as e:
        error_msg = f"Unexpected error connecting to {ip}: {str(e)}"
        print(error_msg)
        import traceback
        print(traceback.format_exc())
        return False, error_msg

    finally:
        if sock:
            try:
                sock.close()
                print(f"Telnet connection to {ip} closed")
            except:
                pass

def connect_to_switch(ip, command, vendor='DCN', username=None, password=None, port=None):
    """
    Establish SSH connection to a network switch (DCN or Huawei) and execute commands.
    
    Args:
        ip (str): IP address of the switch
        command (str): Command to execute on the switch
        vendor (str): 'DCN' or 'HW' for Huawei
        username (str, optional): SSH username (uses vendor default if None)
        password (str, optional): SSH password (uses vendor default if None)
        port (int, optional): SSH port (uses vendor default if None)
        
    Returns:
        tuple: (success: bool, result: str) - Success status and command output or error message
    """
    ssh_client = None
    
    try:
        # Get vendor-specific credentials if not provided
        vendor_config = get_switch_credentials_for_vendor(vendor)
        auth_username = username if username else vendor_config['username']
        auth_password = password if password else vendor_config['password']
        auth_port = port if port else vendor_config['port']
        
        if not auth_username or not auth_password:
            return False, f"Missing credentials for {vendor} switches in environment"
        
        # Initialize SSH client
        ssh_client = paramiko.SSHClient()
        ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        print(f"Connecting to {vendor} switch {ip}:{auth_port} as {auth_username}...")
        
        # For Huawei and older switches, disable stricter algorithms
        # This allows compatibility with switches using older SSH implementations
        connect_kwargs = {
            'hostname': ip,
            'username': auth_username,
            'password': auth_password,
            'port': auth_port,
            'timeout': 30,
            'auth_timeout': 30,
            'banner_timeout': 30,
            'look_for_keys': False,
            'allow_agent': False,
            'compress': False
        }
        
        # For Huawei switches, disable certain algorithms to allow older host key types
        if vendor == 'HW':
            connect_kwargs['disabled_algorithms'] = {
                'pubkeys': ['rsa-sha2-512', 'rsa-sha2-256'],
                'keys': []
            }
        
        # Establish SSH connection
        ssh_client.connect(**connect_kwargs)
        
        print(f"Successfully connected to {vendor} switch {ip}")
        
        # Execute command with timeout
        stdin, stdout, stderr = ssh_client.exec_command(command, timeout=30)
        
        # Wait for command completion
        exit_status = stdout.channel.recv_exit_status()
        
        # Read output
        output = stdout.read().decode('utf-8', errors='ignore')
        error = stderr.read().decode('utf-8', errors='ignore')
        
        print(f"Command executed on {vendor} switch. Exit status: {exit_status}")
        
        # Handle errors
        if exit_status != 0 and error:
            return False, f"Command error (exit {exit_status}): {error}"
        
        return True, output if output else "Command executed successfully with no output"
        
    except paramiko.AuthenticationException as e:
        error_msg = f"Authentication failed for {vendor} switch at {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except paramiko.SSHException as e:
        error_msg = f"SSH error connecting to {vendor} switch at {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except socket.timeout:
        error_msg = f"Connection timeout to {vendor} switch {ip}:{port or 22}"
        print(error_msg)
        return False, error_msg
        
    except socket.error as e:
        error_msg = f"Network error connecting to {vendor} switch {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    except Exception as e:
        error_msg = f"Unexpected error on {vendor} switch {ip}: {str(e)}"
        print(error_msg)
        return False, error_msg
        
    finally:
        if ssh_client:
            try:
                ssh_client.close()
                print(f"SSH connection to {ip} closed")
            except:
                pass

def parse_mikrotik_firewall_output(output):
    """
    Parse MikroTik firewall address-list output into structured data.
    Extracts CSID, IP addresses, and timestamps from RouterOS firewall output.
    Handles both date formats: 'jul/06/2025 15:54:12' and '2025-12-06 15:12:13'
    
    Args:
        output (str): Raw output from MikroTik firewall address-list command
        
    Returns:
        list: List of dictionaries containing parsed firewall entries
              Each entry has: number, csid, ip, date
    """
    lines = output.strip().split('\n')
    extracted_data = []
    current_csid = None
    entry_number = 0
    
    for line in lines:
        # Extract CSID (SI-BK followed by 6 digits)
        csid_match = re.search(r'SI-BK\d{6}', line)
        if csid_match:
            current_csid = csid_match.group()
        
        # Extract IP address
        ip_match = re.search(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', line)
        if ip_match:
            ip = ip_match.group()
            
            # Extract date - handle both formats
            # Format 1: jul/06/2025 15:54:12
            date_match = re.search(r'(\w{3}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})', line)
            if not date_match:
                # Format 2: 2025-12-06 15:12:13
                date_match = re.search(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})', line)
            
            date_added = date_match.group(1) if date_match else 'N/A'
            
            # Create entry
            if current_csid:
                extracted_data.append({
                    "number": str(entry_number),
                    "csid": current_csid,
                    "ip": ip,
                    "date": date_added
                })
                current_csid = None  # Reset after pairing
            else:
                extracted_data.append({
                    "number": str(entry_number),
                    "csid": "N/A",
                    "ip": ip,
                    "date": date_added
                })
            
            entry_number += 1
    
    # Remove duplicates while preserving order
    seen = set()
    unique_data = []
    for item in extracted_data:
        key = (item["csid"], item["ip"])
        if key not in seen:
            seen.add(key)
            unique_data.append(item)
    
    # Sort by CSID (A to Z), putting "N/A" values at the end
    unique_data.sort(key=lambda x: (x["csid"] == "N/A", x["csid"]))
    
    return unique_data



# ============================================================================
#           Device Management Functions (devices.json)
# ============================================================================

def load_devices_active():
    """
    Load active devices from JSON file.
    Reads device configuration from devices-active.json in the same directory.
    
    Returns:
        list: List of active device dictionaries with name, ip, co, location, description
    """
    try:
        # Check if devices.json exists in the same directory as app.py
        devices_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'devices-active.json')
        if os.path.exists(devices_file):
            with open(devices_file, 'r') as f:
                devices = json.load(f)
            return devices
        else:
            # Return empty list if file doesn't exist
            print("Warning: devices-active.json not found. Using empty device list.")
            return []
    except Exception as e:
        print(f"Error loading devices: {e}")
        return []

def load_devices_suspend():
    """
    Load suspended devices from JSON file.
    Reads device configuration from devices-suspend.json in the same directory.
    
    Returns:
        list: List of suspended device dictionaries with name, ip, location, status
    """
    try:
        # Check if devices.json exists in the same directory as app.py
        devices_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'devices-suspend.json')
        if os.path.exists(devices_file):
            with open(devices_file, 'r') as f:
                devices = json.load(f)
            return devices
        else:
            # Return empty list if file doesn't exist
            print("Warning: devices-suspend.json not found. Using empty device list.")
            return []
    except Exception as e:
        print(f"Error loading devices: {e}")
        return []

def load_devices_olt():
    try:
        devices_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'devices-olt.json')
        if os.path.exists(devices_file):
            with open(devices_file, 'r') as f:
                devices = json.load(f)
            return devices
        print("Warning: devices-olt.json not found. Using empty device list.")
        return []
    except Exception as e:
        print(f"Error loading OLT devices: {e}")
        return []

def load_devices_switch():
    """
    Load switch devices from JSON file.
    Reads device configuration from devices-switch.json in the same directory.
    """
    try:
        devices_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'devices-switch.json')
        if os.path.exists(devices_file):
            with open(devices_file, 'r') as f:
                devices = json.load(f)
            return devices
        print("Warning: devices-switch.json not found. Using empty device list.")
        return []
    except Exception as e:
        print(f"Error loading switch devices: {e}")
        return []

# ============================================================================
#           Excel Processing Functions
# ============================================================================

def find_in_all_sheets(session_id, search_value):
    """
    Find search value in all sheets for a specific session.
    Searches through all loaded Excel sheets for the specified value.
    
    Args:
        session_id (str): Session identifier for workbook storage
        search_value (str): Value to search for in the sheets
        
    Returns:
        str: Comma-separated list of sheet names where value was found, or "Not Found"
    """
    if session_id not in workbook_storage:
        return "No data loaded"
    
    workbook_data = workbook_storage[session_id]['data']
    found_sheets = []
    search_str = str(search_value)
    
    for sheet_name, column_data in workbook_data.items():
        str_data = [str(val) for val in column_data if str(val) != 'nan']
        if search_str in str_data:
            found_sheets.append(sheet_name)
    
    if found_sheets:
        return ", ".join(found_sheets)
    else:
        return "Not Found"

# ============================================================================
# Utility Functions
# ============================================================================

def show_startup_message():
    """
    Display startup message for the application.
    Shows application banner and instructions when running as portable version.
    """
    if getattr(sys, 'frozen', False):
        print("\n" + "=" * 60)
        print("           WEB TRACK CSID - PORTABLE VERSION")
        print("=" * 60)
        print("Starting the application...")
        print("Browser will open automatically in 3 seconds...")
        print("\nIf browser doesn't open, manually go to:")
        print("http://localhost:5050")
        print("\nTo stop the application, close this window.")
        print("=" * 60 + "\n")

def open_browser():
    """
    Open web browser automatically after Flask starts.
    Waits 3 seconds for Flask to initialize, then opens the application URL.
    """
    time.sleep(3)
    try:
        webbrowser.open('http://localhost:5050')
        print("✓ Browser opened successfully!")
    except Exception as e:
        print(f"✗ Could not open browser automatically: {e}")
        print("Please manually open: http://localhost:5050")

# ============================================================================
#           Flask Routes - Templates HTML
# ============================================================================

@app.route('/')
def index():
    """Main page"""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    return render_template('home.html')

@app.route('/manage-active.html')
def manage_active():
    """Route for the Manage Active - Batch CSID Status Checker page"""
    return render_template('manage-active.html')

@app.route('/tracking-co.html')
def track_home():
    return render_template('tracking-co.html')

@app.route('/tracking-odp.html')
def track_odp():
    return render_template('tracking-odp.html')
    
@app.route('/tracking-ip.html')
def track_ip():
    return render_template('tracking-ip.html')

@app.route('/automate-suspend.html')
def automate_suspend():
    """Route for the SUSPEND CO automation page with MikroTik integration"""
    return render_template('automate-suspend.html')

@app.route('/simple-queue-parser.html')
def simple_queue_parser():
    """Route for the Simple Queue Parser page"""
    return render_template('simple-queue-parser.html')

@app.route('/parse-mac-olt.html')
def parse_mac_olt():
    """Route for the Parse MAC OLT page"""
    return render_template('parse-mac-olt.html')

@app.route('/all-olt-parser.html')
def all_olt_parser():   
    """Route for the All OLT Parser page"""         
    return render_template('all-olt-parser.html')

@app.route('/custom-command.html')
def custom_command():
    """Route for the Custom Command Executor page"""
    return render_template('custom-command.html')

@app.route('/custom-olt-command.html')
def custom_olt_command():
    return render_template('custom-olt-command.html')

@app.route('/custom-switch-command.html')
def custom_switch_command():
    """Route for the Custom Switch Command management page"""
    return render_template('custom-switch-command.html')

@app.route('/address-list-parser.html')
def address_list_parser():
    """Route for the Address List Parser page"""
    return render_template('address-list-parser.html')

@app.route('/arp-parser.html')
def arp_parser():
    """Route for the ARP Parser page"""
    return render_template('arp-parser.html')
    
#           Flask Routes - SCRIPT TRACKING EXCEL
# ============================================================================

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle Excel file upload"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Read Excel file
        workbook_data = {}
        with pd.ExcelFile(file) as xls:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if len(df.columns) >= 3:  # Check if column C exists
                    column_c = df.iloc[:, 2].astype(str).dropna()  # Column C (index 2)
                    workbook_data[sheet_name] = column_c.tolist()
                else:
                    workbook_data[sheet_name] = []
        
        # Store in session storage
        session_id = session['session_id']
        workbook_storage[session_id] = {
            'data': workbook_data,
            'filename': secure_filename(file.filename),
            'upload_time': datetime.now().isoformat()
        }
        
        return jsonify({
            'success': True,
            'message': f'Loaded data from {len(workbook_data)} sheets',
            'sheets': list(workbook_data.keys())
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to load Excel file: {str(e)}'}), 500

@app.route('/export', methods=['POST'])
def export_data():
    """Export data to Excel"""
    data = request.get_json()
    export_data = data.get('data', [])
    
    if not export_data:
        return jsonify({'error': 'No data to export'}), 400
    
    try:
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='CSID_Results')
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'csid_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500

@app.route('/search-odp', methods=['POST'])
def search_odp():
    odp_id = request.form.get('odp_id')
    file = request.files['file']
    
    wb = openpyxl.load_workbook(file)
    results = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "TRACK ODP":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() == odp_id.strip():
                results.append({
                    'ip': row[1],
                    'csid': row[2]
                })
    
    return jsonify(results)

@app.route('/batch-search-odp', methods=['POST'])
def batch_search_odp():
    """Batch search for multiple ODP IDs"""
    try:
        batch_odp_ids_json = request.form.get('batch_odp_ids')
        file = request.files['file']
        
        if not batch_odp_ids_json or not file:
            return jsonify({'error': 'Missing required data'}), 400
        
        # Parse the JSON string to get the list of ODP IDs
        batch_odp_ids = json.loads(batch_odp_ids_json)
        
        if not isinstance(batch_odp_ids, list) or len(batch_odp_ids) == 0:
            return jsonify({'error': 'Invalid ODP IDs format'}), 400
        
        wb = openpyxl.load_workbook(file)
        all_results = []
        found_odp_ids = set()
        
        # Search for each ODP ID
        for odp_id in batch_odp_ids:
            odp_id = str(odp_id).strip()
            if not odp_id:
                continue
                
            # Search in all sheets except "TRACK ODP"
            for sheet_name in wb.sheetnames:
                if sheet_name == "TRACK ODP":
                    continue
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]).strip() == odp_id:
                        all_results.append({
                            'odp_id': odp_id,
                            'ip': row[1] if row[1] else 'N/A',
                            'csid': row[2] if row[2] else 'N/A'
                        })
                        found_odp_ids.add(odp_id)
        
        # Create summary
        summary = {
            'total': len(batch_odp_ids),
            'found': len(found_odp_ids),
            'not_found': len(batch_odp_ids) - len(found_odp_ids)
        }
        
        return jsonify({
            'results': all_results,
            'summary': summary
        })
        
    except json.JSONDecodeError:
        return jsonify({'error': 'Invalid JSON format for ODP IDs'}), 400
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/search-ip', methods=['POST'])
def search_ip():
    csid_input = request.form.get('csid')
    file = request.files['file']
    
    # Parse multiple CSIDs - split by newlines, commas, or semicolons
    csids = []
    if csid_input:
        # Split by various delimiters and clean up
        csid_list = re.split(r'[,;\n\r]+', csid_input.strip())
        csids = [csid.strip() for csid in csid_list if csid.strip()]
    
    if not csids:
        return jsonify({'error': 'Please provide at least one CSID'})
    
    wb = openpyxl.load_workbook(file)
    
    # Build the lookup table first by scanning all sheets once
    ip_map = {}  # Dictionary: csid -> {'ip': ip_address, 'sheet': sheet_name}
    
    # Single pass through all sheets to build the lookup map
    for sheet_name in wb.sheetnames:
        if sheet_name == "TRACK ODP":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 2 and row[2]:  # Check if CSID column has value
                row_csid = str(row[2]).strip()
                if row_csid:  # Only add non-empty CSIDs
                    ip_map[row_csid] = {
                        'ip': row[1] if len(row) > 1 and row[1] else 'N/A',
                        'sheet': sheet_name
                    }
    
    # Fast O(1) lookup for each searched CSID
    results = []
    found_csids = set()
    
    for search_csid in csids:
        if search_csid in ip_map:
            # Found the CSID
            data = ip_map[search_csid]
            results.append({
                'csid': search_csid,
                'ip': data['ip'],
                'sheet': data['sheet']
            })
            found_csids.add(search_csid)
    
    # Add entries for CSIDs that were not found
    not_found_csids = set(csids) - found_csids
    for csid in not_found_csids:
        results.append({
            'csid': csid,
            'ip': 'Not Found',
            'sheet': 'N/A'
        })
    
    # Sort results to show found ones first, then not found
    results.sort(key=lambda x: (x['ip'] == 'Not Found', x['csid']))
    
    return jsonify(results)

@app.route('/find-odp', methods=['POST'])
def find_odp():
    csid_input = request.form.get('csid')
    file = request.files['file']
    
    # Parse multiple CSIDs - split by newlines, commas, or semicolons
    csids = []
    if csid_input:
        # Split by various delimiters and clean up
        csid_list = re.split(r'[,;\n\r]+', csid_input.strip())
        csids = [csid.strip() for csid in csid_list if csid.strip()]
    
    if not csids:
        return jsonify({'error': 'Please provide at least one CSID'})
    
    wb = openpyxl.load_workbook(file)
    
    # Build the lookup table first by scanning all sheets once
    odp_map = {}  # Dictionary: csid -> {'odp': odp_id, 'sheet': sheet_name}
    
    # Single pass through all sheets to build the lookup map
    for sheet_name in wb.sheetnames:
        if sheet_name == "TRACK ODP":
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 2 and row[2]:  # Check if CSID column has value
                row_csid = str(row[2]).strip()
                if row_csid:  # Only add non-empty CSIDs
                    odp_map[row_csid] = {
                        'odp': row[0] if len(row) > 0 and row[0] else 'N/A',
                        'sheet': sheet_name
                    }
    
    # Fast O(1) lookup for each searched CSID
    results = []
    found_csids = set()
    
    for search_csid in csids:
        if search_csid in odp_map:
            # Found the CSID
            data = odp_map[search_csid]
            results.append({
                'csid': search_csid,
                'odp': data['odp'],
                'sheet': data['sheet']
            })
            found_csids.add(search_csid)
    
    # Add entries for CSIDs that were not found
    not_found_csids = set(csids) - found_csids
    for csid in not_found_csids:
        results.append({
            'csid': csid,
            'odp': 'Not Found',
            'sheet': 'N/A'
        })
    
    # Sort results to show found ones first, then not found
    results.sort(key=lambda x: (x['odp'] == 'Not Found', x['csid']))
    
    return jsonify(results)

# ============================================================================
#       Flask Routes - PARSING CSID LOGIC
# ============================================================================

@app.route('/add_csid', methods=['POST'])
def add_csid():
    """Add single CSID"""
    data = request.get_json()
    csid = data.get('csid', '').strip()
    
    if not csid:
        return jsonify({'error': 'CSID cannot be empty'}), 400
    
    session_id = session['session_id']
    found_sheets = find_in_all_sheets(session_id, csid)
    
    return jsonify({
        'success': True,
        'csid': csid,
        'found_sheets': found_sheets
    })

@app.route('/add_bulk_csids', methods=['POST'])
def add_bulk_csids():
    """Add multiple CSIDs"""
    data = request.get_json()
    csids_text = data.get('csids', '').strip()
    
    if not csids_text:
        return jsonify({'error': 'No CSIDs provided'}), 400
    
    # Parse CSIDs
    csids = re.split(r'[\s,;\n\t]+', csids_text)
    csids = [csid.strip() for csid in csids if csid.strip()]
    
    if not csids:
        return jsonify({'error': 'No valid CSIDs found'}), 400
    
    session_id = session['session_id']
    results = []
    
    for csid in csids:
        found_sheets = find_in_all_sheets(session_id, csid)
        results.append({
            'csid': csid,
            'found_sheets': found_sheets
        })
    
    return jsonify({
        'success': True,
        'results': results,
        'count': len(results)
    })

@app.route('/refresh_csids', methods=['POST'])
def refresh_csids():
    """Refresh existing CSIDs"""
    data = request.get_json()
    csids = data.get('csids', [])
    
    if not csids:
        return jsonify({'error': 'No CSIDs to refresh'}), 400
    
    session_id = session['session_id']
    results = []
    
    for csid in csids:
        found_sheets = find_in_all_sheets(session_id, csid)
        results.append({
            'csid': csid,
            'found_sheets': found_sheets
        })
    
    return jsonify({
        'success': True,
        'results': results
    })

@app.route('/extract-suspend', methods=['POST'])
def extract_suspend():
    """Extract CSID and IP addresses from SUSPEND CO text"""
    input_text = request.form.get('input_text', '')
    
    lines = input_text.strip().split('\n')
    extracted_data = []
    current_csid = None
    
    for line in lines:
        # Extract CSID (SI-BK followed by 6 digits)
        csid_match = re.search(r'SI-BK\d{6}', line)
        if csid_match:
            current_csid = csid_match.group()
        
        # Extract IP address
        ip_match = re.search(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', line)
        if ip_match:
            ip = ip_match.group()
            if current_csid:
                extracted_data.append({"csid": current_csid, "ip": ip})
                current_csid = None  # Reset after pairing
            else:
                extracted_data.append({"csid": "N/A", "ip": ip})
    
    # Remove duplicates while preserving order
    seen = set()
    unique_data = []
    for item in extracted_data:
        key = (item["csid"], item["ip"])
        if key not in seen:
            seen.add(key)
            unique_data.append(item)
    
    # Sort by CSID (A to Z), putting "N/A" values at the end
    unique_data.sort(key=lambda x: (x["csid"] == "N/A", x["csid"]))
    
    return jsonify({"results": unique_data, "count": len(unique_data)})

# ============================================================================
#       Flask Routes - MANAGEMENT SESSION EXCEL
# ============================================================================

@app.route('/status')
def get_status():
    """Get current session status"""
    session_id = session.get('session_id')
    
    if session_id in workbook_storage:
        wb_info = workbook_storage[session_id]
        return jsonify({
            'loaded': True,
            'filename': wb_info['filename'],
            'sheets': list(wb_info['data'].keys()),
            'upload_time': wb_info['upload_time']
        })
    else:
        return jsonify({'loaded': False})

@app.route('/reset_file', methods=['POST'])
def reset_file():
    """Reset the uploaded file data"""
    session_id = session.get('session_id')
    if session_id in workbook_storage:
        del workbook_storage[session_id]
    return jsonify({'success': True, 'message': 'File data reset successfully'})

# ============================================================================
#           Flask Routes - ALL API MIKROTIK SCRIPT
# ============================================================================

@app.route('/api/mikrotik/firewall-list', methods=['POST'])
def get_firewall_list():
    """
    API endpoint to get firewall address list from MikroTik device
    Now accepts credentials from frontend
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        ip = data.get('ip')
        command = data.get('command', '/ip firewall address-list print where list="SUSPEND_SOLNET"')
        username = data.get('username')
        password = data.get('password')
        
        if not ip:
            return jsonify({'error': 'IP address is required'}), 400
        
        # Connect to MikroTik and execute command with provided credentials
        success, output = connect_to_mikrotik(ip, command, username, password)
        
        if not success:
            return jsonify({'error': output}), 500
        
        # Parse the output
        results = parse_mikrotik_firewall_output(output)
        
        return jsonify({
            'success': True,
            'device_ip': ip,
            'results': results,
            'raw_output': output  # Include raw output for debugging
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/mikrotik/test-connection', methods=['POST'])
def test_connection():
    """
    Test connection to MikroTik device
    """
    try:
        data = request.get_json()
        ip = data.get('ip')
        username = data.get('username')
        password = data.get('password')
        
        if not ip:
            return jsonify({'error': 'IP address is required'}), 400
        
        # Test connection with a simple command
        success, output = connect_to_mikrotik(ip, '/system identity print', username, password)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Connection successful',
                'device_ip': ip,
                'identity': output.strip()
            })
        else:
            return jsonify({
                'success': False,
                'error': output
            }), 500
            
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/mikrotik/list-names', methods=['POST'])
def get_list_names():
    """
    Get all available address-list names from MikroTik
    """
    try:
        data = request.get_json()
        ip = data.get('ip')
        username = data.get('username')
        password = data.get('password')
        
        if not ip:
            return jsonify({'error': 'IP address is required'}), 400
        
        # Get all address-lists to see available list names
        success, output = connect_to_mikrotik(ip, '/ip firewall address-list print', username, password)
        
        if not success:
            return jsonify({'error': output}), 500
        
        # Extract unique list names
        lines = output.split('\n')
        list_names = set()
        
        for line in lines:
            # Look for list= patterns
            list_match = re.search(r'list=([^\s]+)', line)
            if list_match:
                list_name = list_match.group(1).strip('"')
                list_names.add(list_name)
        
        return jsonify({
            'success': True,
            'device_ip': ip,
            'available_lists': sorted(list(list_names)),
            'raw_output': output
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/mikrotik/debug-output', methods=['POST'])
def debug_output():
    """
    Debug endpoint to see raw MikroTik output
    """
    try:
        data = request.get_json()
        ip = data.get('ip')
        command = data.get('command', '/ip firewall address-list print where list="SUSPEND_SOLNET"')
        username = data.get('username')
        password = data.get('password')
        
        if not ip:
            return jsonify({'error': 'IP address is required'}), 400
        
        # Connect to MikroTik and execute command
        success, output = connect_to_mikrotik(ip, command, username, password)
        
        if not success:
            return jsonify({'error': output}), 500
        
        # Return raw output split by lines for easier debugging
        lines = output.split('\n')
        numbered_lines = []
        for i, line in enumerate(lines):
            numbered_lines.append(f"{i:2d}: {repr(line)}")
        
        return jsonify({
            'success': True,
            'device_ip': ip,
            'raw_output': output,
            'lines': numbered_lines
        })
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500
        
@app.route('/api/mikrotik/execute-custom-command', methods=['POST'])
def execute_custom_command():
    """
    API endpoint to execute custom MikroTik commands on a single device
    This endpoint is designed to work with the custom command executor frontend
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        ip = data.get('ip')
        command = data.get('command')
        username = data.get('username')
        password = data.get('password')
        
        # Validate required parameters
        if not ip:
            return jsonify({'error': 'IP address is required'}), 400
        
        if not command:
            return jsonify({'error': 'Command is required'}), 400
        
        if not username:
            return jsonify({'error': 'Username is required'}), 400
        
        # Handle stored credentials - if password is null, look up stored password
        if password is None:
            stored_credentials = load_mikrotik_credentials()
            user_credential = next((cred for cred in stored_credentials if cred['username'] == username), None)
            if user_credential and user_credential.get('password'):
                password = user_credential['password']
            else:
                return jsonify({'error': 'No stored password found for this user'}), 400
        
        if not password:
            return jsonify({'error': 'Password is required'}), 400
        
        # Execute the command on the MikroTik device
        success, output = connect_to_mikrotik(ip, command, username, password)
        
        if success:
            return jsonify({
                'success': True,
                'device_ip': ip,
                'command': command,
                'output': output,
                'timestamp': datetime.now().isoformat()
            })
        else:
            return jsonify({
                'success': False,
                'device_ip': ip,
                'command': command,
                'error': output,
                'timestamp': datetime.now().isoformat()
            }), 500
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/api/olt/execute-custom-command', methods=['POST'])
def execute_olt_custom_command():
    try:
        data = request.get_json()

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        ip = data.get('ip')
        command = data.get('command')
        username = data.get('username')
        password = data.get('password')

        if not ip:
            return jsonify({'error': 'IP address is required'}), 400

        if not command:
            return jsonify({'error': 'Command is required'}), 400

        success, output = connect_to_olt(ip, command, username, password)

        if success:
            return jsonify({
                'success': True,
                'device_ip': ip,
                'command': command,
                'output': output,
                'timestamp': datetime.now().isoformat()
            })

        return jsonify({
            'success': False,
            'device_ip': ip,
            'command': command,
            'error': output,
            'timestamp': datetime.now().isoformat()
        }), 500

    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500
        
@app.route('/api/mikrotik/batch-status-check', methods=['POST'])
def batch_status_check():
    try:
        data = request.get_json()
        
        # Debug logging
        app.logger.info(f"Received data: {data}")
        
        if not data or 'csids' not in data:
            return jsonify({'error': 'No CSIDs provided'}), 400
        
        csid_items = data['csids']  # Note: Using the same parameter name as frontend
        username = data.get('username')
        password = data.get('password')
        results = []
        
        # Group by device IP
        device_groups = {}
        for item in csid_items:
            ip = item.get('deviceIp')
            if ip and ip != 'Unknown':
                if ip not in device_groups:
                    device_groups[ip] = []
                device_groups[ip].append(item)
        
        # Process each device group
        for ip, items in device_groups.items():
            try:
                # First check suspended list (SUSPEND_SOLNET)
                suspend_success, suspend_output = connect_to_mikrotik(
                    ip,
                    '/ip firewall address-list print where list="SUSPEND_SOLNET"',
                    username=username,
                    password=password
                )
                
                suspended_csids = []
                if suspend_success:
                    suspended_data = parse_mikrotik_firewall_output(suspend_output)
                    suspended_csids = [item['csid'] for item in suspended_data if item['csid'] != 'N/A']
                
                # Then check main address list with detailed info
                main_success, main_output = connect_to_mikrotik(
                    ip,
                    '/ip firewall address-list print detail',
                    username=username,
                    password=password
                )
                
                address_list_csids = {}
                if main_success:
                    # Parse the detailed output to get enabled/disabled status
                    lines = main_output.strip().split('\n')
                    
                    current_csid = None
                    is_disabled = False
                    
                    for line in lines:
                        line_stripped = line.strip()
                        
                        # Skip empty lines
                        if not line_stripped:
                            continue
                        
                        # Look for CSID in current line
                        csid_match = re.search(r'SI-BK\d{6}', line_stripped)
                        if csid_match:
                            current_csid = csid_match.group()
                        
                        # Check if this is a status line (starts with number and optionally X)
                        # Format examples:
                        # "25 X ;;; SUDAH BONGKAR by Wawan"
                        # "26   ;;; Some comment"
                        # "27 X"
                        status_match = re.match(r'^\s*(\d+)\s*(X)?\s*(?:;;;.*)?$', line_stripped)
                        if status_match:
                            x_flag = status_match.group(2)
                            is_disabled = bool(x_flag)  # True if X is present, False otherwise
                        
                        # Check for explicit disabled status in any line
                        if 'disabled=yes' in line.lower():
                            is_disabled = True
                        elif 'disabled=no' in line.lower():
                            is_disabled = False
                        
                        # If we found a CSID and we're processing lines that belong to this entry
                        # Look for IP address to confirm this is a complete entry
                        ip_match = re.search(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', line_stripped)
                        if current_csid and ip_match:
                            # Save the status for this CSID
                            address_list_csids[current_csid] = 'disabled' if is_disabled else 'enabled'
                            app.logger.info(f"Found CSID {current_csid} with status: {'disabled' if is_disabled else 'enabled'}")
                            # Reset for next entry
                            current_csid = None
                            is_disabled = False
                
                # Determine status for each CSID based on new logic
                for item in items:
                    csid = item['csid']
                    
                    # Step 1: Check if in SUSPEND_SOLNET
                    if csid in suspended_csids:
                        status = 'SUSPENDED'
                    # Step 2: Check if in address-list
                    elif csid in address_list_csids:
                        if address_list_csids[csid] == 'enabled':
                            status = 'ACTIVE'
                        else:
                            status = 'DISABLED'
                    # Step 3: Not found in either
                    else:
                        status = 'NOT_FOUND'
                    
                    app.logger.info(f"CSID {csid}: Final status = {status}")
                    
                    results.append({
                        'csid': csid,
                        'co': item['co'],
                        'deviceIp': ip,
                        'status': status,
                        'error': None
                    })
                    
            except Exception as e:
                app.logger.error(f"Error processing {ip}: {str(e)}")
                for item in items:
                    results.append({
                        'csid': item['csid'],
                        'co': item['co'],
                        'deviceIp': ip,
                        'status': 'NOT_FOUND',
                        'error': str(e)
                    })
        
        # Calculate statistics
        stats = {
            'active': len([r for r in results if r['status'] == 'ACTIVE']),
            'suspended': len([r for r in results if r['status'] == 'SUSPENDED']),
            'disabled': len([r for r in results if r['status'] == 'DISABLED']),
            'not_found': len([r for r in results if r['status'] == 'NOT_FOUND']),
            'total': len(results)
        }
        
        return jsonify({
            'success': True,
            'results': results,
            'stats': stats
        })
        
    except Exception as e:
        app.logger.error(f"Server error: {str(e)}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500

# ============================================================================
#           Flask Routes - Device API Endpoints (ALL ROUTE API MIKROTIK)
# ============================================================================

@app.route('/api/devices/co-mapping', methods=['GET'])
def get_co_mapping():
    """
    Get CO to IP mapping for the frontend
    """
    try:
        devices = load_devices_active()
        co_mapping = {}
        
        for device in devices:
            if device.get('co') and device.get('ip'):
                co_mapping[device['co']] = device['ip']
        
        return jsonify({
            'success': True,
            'co_mapping': co_mapping
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to load CO mapping: {str(e)}'}), 500

@app.route('/api/devices-suspend', methods=['GET'])
def get_devices_suspend():
    """
    Get list of available MikroTik devices from JSON file
    """
    devices = load_devices_suspend()
    return jsonify({'devices': devices})

@app.route('/api/devices-olt', methods=['GET'])
def get_devices_olt():
    devices = load_devices_olt()
    return jsonify({'devices': devices})

@app.route('/api/devices-switch', methods=['GET'])
def get_devices_switch():
    """Get list of available switch devices from JSON file"""
    devices = load_devices_switch()
    return jsonify({'devices': devices})

@app.route('/api/devices-active', methods=['GET'])
def get_devices_active():
    """
    Get list of available MikroTik devices from JSON file
    """
    devices = load_devices_active() 
    return jsonify({'devices': devices})

@app.route('/api/mikrotik-credentials', methods=['GET'])
def get_mikrotik_credentials():
    """
    Get available MikroTik users from environment variables
    """
    try:
        credentials = []
        
        # Read all available users (up to 10 for safety)
        for i in range(1, 11):
            if i == 1:
                username_key = 'MIKROTIK_USERNAME'
                password_key = 'MIKROTIK_PASSWORD'
            else:
                username_key = f'MIKROTIK_USERNAME_{i}'
                password_key = f'MIKROTIK_PASSWORD_{i}'
            
            username = os.getenv(username_key)
            password = os.getenv(password_key)
            
            if username:
                credentials.append({
                    'username': username,
                    'hasPassword': bool(password and password.strip())
                })
        
        return jsonify({
            'success': True,
            'credentials': credentials
        })
        
    except Exception as e:
        return jsonify({'error': f'Failed to load credentials: {str(e)}'}), 500


def infer_switch_vendor(device_name):
    """Infer switch vendor from the device name or vendor label."""
    if not device_name:
        return 'DCN'

    name = str(device_name).upper()
    if 'HW' in name or 'HUAWEI' in name:
        return 'HW'
    if 'DCN' in name:
        return 'DCN'
    return 'DCN'


def get_switch_credentials_for_vendor(vendor_name):
    """Return the credentials for a given vendor from environment variables."""
    vendor = (vendor_name or 'DCN').upper()

    if vendor == 'HW':
        username = os.getenv('SWITCH_HW_USERNAME')
        password = os.getenv('SWITCH_HW_PASSWORD')
        port = int(os.getenv('SWITCH_HW_PORT', '22'))
    else:
        username = os.getenv('SWITCH_DCN_USERNAME')
        password = os.getenv('SWITCH_DCN_PASSWORD')
        port = int(os.getenv('SWITCH_DCN_PORT', '22'))

    return {
        'vendor': vendor,
        'username': username,
        'password': password,
        'port': port,
        'hasPassword': bool(password and password.strip())
    }


@app.route('/api/switch-credentials', methods=['GET'])
def get_switch_credentials():
    """Get available switch credentials by vendor."""
    try:
        vendors = ['DCN', 'HW']
        credentials = []

        for vendor in vendors:
            config = get_switch_credentials_for_vendor(vendor)
            if config['username']:
                credentials.append({
                    'vendor': config['vendor'],
                    'username': config['username'],
                    'password': config['password'] if config['password'] and config['password'].strip() else None,
                    'hasPassword': config['hasPassword'],
                    'port': config['port']
                })

        return jsonify({'success': True, 'credentials': credentials})
    except Exception as e:
        return jsonify({'error': f'Failed to load switch credentials: {str(e)}'}), 500


@app.route('/api/switch/execute-command', methods=['POST'])
def execute_switch_command():
    """Execute a command on a network switch (DCN or Huawei)."""
    try:
        # Parse JSON with error handling
        data = request.get_json(force=True, silent=False)
        if not data:
            return jsonify({'success': False, 'error': 'No JSON data provided'}), 400
        
        # Safely handle potential None values
        ip = (data.get('ip') or '').strip()
        command = (data.get('command') or '').strip()
        vendor = (data.get('vendor') or 'DCN').upper()
        username = (data.get('username') or '').strip() or None
        password = (data.get('password') or '').strip() or None
        
        # Validate required fields
        if not ip:
            return jsonify({'success': False, 'error': 'IP address is required'}), 400
        if not command:
            return jsonify({'success': False, 'error': 'Command is required'}), 400
        
        print(f"Executing command on {vendor} switch {ip}: {command[:50]}...")
        
        # Execute command on switch
        success, output = connect_to_switch(
            ip=ip,
            command=command,
            vendor=vendor,
            username=username,
            password=password
        )
        
        # Ensure output is a string
        output_str = str(output) if output else ''
        
        return jsonify({
            'success': success,
            'output': output_str,
            'error': None if success else output_str
        })
        
    except Exception as e:
        error_msg = f"Error executing switch command: {str(e)}"
        print(f"EXCEPTION in /api/switch/execute-command: {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': error_msg}), 500

# ============================================================================
# Application Entry Point
# ============================================================================

if __name__ == '__main__':
    print("Starting WEB TRACK CSID with MikroTik Integration...")
    print("Make sure to:")
    print("1. Install required packages: pip install flask paramiko flask-cors pandas openpyxl")
    print("2. Create devices.json file in the same directory as app.py")
    print("3. Ensure your MikroTik devices have SSH enabled")
    print("4. Update your automate-suspend.html template to include MikroTik functionality")
    
    show_startup_message()
    
    # Start browser in background
    if getattr(sys, 'frozen', False):
        threading.Thread(target=open_browser, daemon=True).start()
    
    app.run(debug=False, host=os.getenv('FLASK_HOST'), port=int(os.getenv('FLASK_PORT')))

